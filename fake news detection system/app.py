"""
Flask Web Application for Fake News Detection System
Provides web interface for users to verify news authenticity
"""

from flask import Flask, render_template, request, jsonify, send_file
import os
import sys

if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

import pickle
from datetime import datetime
from io import BytesIO

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from utils.preprocessing import TextPreprocessor
from utils.feature_extraction import FeatureExtractor
from utils.helpers import format_prediction, validate_text_input
import database  # Import database module

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here-change-in-production'

# Global variables for model and vectorizer
model = None
vectorizer = None
preprocessor = None


def load_models():
    """Load trained model and vectorizer"""
    global model, vectorizer, preprocessor
    
    try:
        # Load best model
        model_path = 'models/best_model.pkl'
        if os.path.exists(model_path):
            with open(model_path, 'rb') as f:
                model_data = pickle.load(f)
                # Handle both dictionary and direct model formats
                if isinstance(model_data, dict):
                    model = model_data['model']
                else:
                    model = model_data
            print("✓ Model loaded successfully")
        else:
            print("⚠ Warning: Model file not found. Please run train.py first.")
        
        # Load vectorizer
        vectorizer_path = 'models/tfidf_vectorizer.pkl'
        if os.path.exists(vectorizer_path):
            with open(vectorizer_path, 'rb') as f:
                vectorizer = pickle.load(f)
            print("✓ Vectorizer loaded successfully")
        else:
            print("⚠ Warning: Vectorizer file not found. Please run train.py first.")
        
        # Initialize preprocessor
        preprocessor = TextPreprocessor(use_lemmatization=True)
        print("✓ Preprocessor initialized")
        
    except Exception as e:
        print(f"✗ Error loading models: {str(e)}")


@app.route('/')
def index():
    """Home page"""
    return render_template('index.html')


@app.route('/predict', methods=['POST'])
def predict():
    """
    Predict whether news is fake or real
    
    Returns:
        JSON response with prediction results
    """
    try:
        # Check if models are loaded
        if model is None or vectorizer is None or preprocessor is None:
            return jsonify({
                'error': 'Models not loaded. Please run train.py first to train the models.'
            }), 500
        
        # Get text from request
        data = request.get_json()
        text = data.get('text', '')
        
        # Validate input
        is_valid, error_message = validate_text_input(text)
        if not is_valid:
            return jsonify({'error': error_message}), 400
        
        # Preprocess text
        processed_text = preprocessor.preprocess(text)
        
        # Extract features
        features = vectorizer.transform([processed_text])
        
        # Make prediction
        prediction = model.predict(features)[0]
        probabilities = model.predict_proba(features)[0]
        
        # Get confidence (probability of predicted class)
        confidence = probabilities[prediction]
        
        # Format result
        result = format_prediction(prediction, confidence)
        result['text_length'] = len(text)
        result['processed_text_sample'] = processed_text[:100] + '...' if len(processed_text) > 100 else processed_text
        
        # Save to database
        try:
            ip_address = request.remote_addr
            database.save_prediction(
                text=text,
                prediction=result['label'],
                confidence=result['confidence'],
                message=result['message'],
                ip_address=ip_address
            )
        except Exception as db_error:
            print(f"Warning: Failed to save to database: {db_error}")
        
        return jsonify(result)
    
    except Exception as e:
        return jsonify({'error': f'Prediction error: {str(e)}'}), 500


@app.route('/about')
def about():
    """About page"""
    return render_template('about.html')


@app.route('/admin')
def admin():
    """Admin dashboard page"""
    return render_template('admin.html')


@app.route('/api/health', methods=['GET'])
def health():
    """Health check endpoint"""
    status = {
        'status': 'running',
        'model_loaded': model is not None,
        'vectorizer_loaded': vectorizer is not None,
        'preprocessor_loaded': preprocessor is not None
    }
    return jsonify(status)


@app.route('/export/pdf', methods=['POST'])
def export_pdf():
    """Export prediction results to PDF"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        
        data = request.get_json()
        text = data.get('text', '')
        label = data.get('label', '')
        confidence = data.get('confidence', 0)
        message = data.get('message', '')
        
        # Create PDF in memory
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#6366f1'),
            spaceAfter=30,
        )
        story.append(Paragraph("Fake News Detection Report", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        story.append(Paragraph(f"<b>Generated:</b> {timestamp}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Results table
        result_data = [
            ['Prediction Result', label],
            ['Confidence Level', f'{confidence}%'],
            ['Analysis', message]
        ]
        
        result_table = Table(result_data, colWidths=[2*inch, 4*inch])
        result_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        story.append(result_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Input text
        story.append(Paragraph("<b>Analyzed Text:</b>", styles['Heading2']))
        story.append(Spacer(1, 0.1*inch))
        
        text_style = ParagraphStyle(
            'CustomBody',
            parent=styles['Normal'],
            fontSize=10,
            leading=14,
        )
        story.append(Paragraph(text.replace('\n', '<br/>'), text_style))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'fake_news_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate PDF: {str(e)}'}), 500


@app.route('/export/excel', methods=['POST'])
def export_excel():
    """Export prediction results to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        data = request.get_json()
        text = data.get('text', '')
        label = data.get('label', '')
        confidence = data.get('confidence', 0)
        message = data.get('message', '')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Fake News Detection Report"
        
        # Header styling
        header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        
        # Title
        ws['A1'] = "Fake News Detection Report"
        ws['A1'].font = Font(bold=True, size=16, color="6366f1")
        ws.merge_cells('A1:B1')
        
        # Timestamp
        ws['A2'] = "Generated:"
        ws['B2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A2'].font = Font(bold=True)
        
        # Results section
        ws['A4'] = "Prediction Result"
        ws['B4'] = label
        ws['A4'].font = Font(bold=True)
        ws['A4'].fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")
        
        ws['A5'] = "Confidence Level"
        ws['B5'] = f"{confidence}%"
        ws['A5'].font = Font(bold=True)
        ws['A5'].fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")
        
        ws['A6'] = "Analysis"
        ws['B6'] = message
        ws['A6'].font = Font(bold=True)
        ws['A6'].fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")
        
        # Analyzed text
        ws['A8'] = "Analyzed Text:"
        ws['A8'].font = Font(bold=True, size=12)
        ws['A9'] = text
        ws['A9'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 80
        
        # Row heights
        ws.row_dimensions[9].height = max(15 * (len(text) // 80 + 1), 30)
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'fake_news_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate Excel: {str(e)}'}), 500


@app.route('/api/statistics', methods=['GET'])
def get_statistics():
    """Get prediction statistics from database"""
    try:
        stats = database.get_statistics()
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/history', methods=['GET'])
def get_history():
    """Get recent prediction history"""
    try:
        limit = request.args.get('limit', 50, type=int)
        history = database.get_recent_predictions(limit)
        
        # Convert to list of dicts
        history_list = []
        for row in history:
            history_list.append({
                'id': row['id'],
                'text_preview': row['text_preview'],
                'prediction': row['prediction'],
                'confidence': row['confidence'],
                'timestamp': row['timestamp']
            })
        
        return jsonify({'history': history_list, 'count': len(history_list)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/prediction/<int:prediction_id>', methods=['GET'])
def get_single_prediction(prediction_id):
    """Get a single prediction by ID for viewing full text"""
    try:
        prediction = database.get_prediction_by_id(prediction_id)
        if prediction:
            return jsonify({
                'id': prediction['id'],
                'text': prediction['text'],
                'prediction': prediction['prediction'],
                'confidence': prediction['confidence'],
                'message': prediction['message'],
                'timestamp': prediction['timestamp']
            })
        else:
            return jsonify({'error': 'Prediction not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/clear-history', methods=['POST'])
def clear_history():
    """Clear all prediction history"""
    try:
        deleted_count = database.delete_all_predictions()
        return jsonify({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Successfully deleted {deleted_count} predictions'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export-all', methods=['GET'])
def export_all_data():
    """Export all prediction data to CSV"""
    try:
        csv_data = database.export_to_csv()
        
        buffer = BytesIO()
        buffer.write(csv_data.encode('utf-8'))
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'all_predictions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export-all-pdf', methods=['GET'])
def export_all_data_pdf():
    """Export all prediction history to PDF"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        
        # Get all predictions
        predictions = database.get_all_predictions(limit=1000)
        
        if not predictions:
            return jsonify({'error': 'No predictions found'}), 404
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, 
                              rightMargin=0.5*inch, leftMargin=0.5*inch,
                              topMargin=0.75*inch, bottomMargin=0.75*inch)
        
        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#6366f1'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        title = Paragraph("📊 Prediction History Report", title_style)
        elements.append(title)
        
        # Summary statistics
        stats = database.get_statistics()
        summary_style = ParagraphStyle(
            'Summary',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=20,
            alignment=TA_LEFT
        )
        
        summary_text = f"""
        <b>Total Predictions:</b> {stats['total_predictions']}<br/>
        <b>Fake News Detected:</b> {stats['fake_count']}<br/>
        <b>Real News Detected:</b> {stats['real_count']}<br/>
        <b>Average Confidence:</b> {stats['average_confidence']}%<br/>
        <b>Generated:</b> {datetime.now().strftime('%B %d, %Y at %I:%M:%S %p')}
        """
        summary = Paragraph(summary_text, summary_style)
        elements.append(summary)
        elements.append(Spacer(1, 0.3*inch))
        
        # Table header
        table_data = [['ID', 'Text Preview', 'Prediction', 'Confidence', 'Timestamp']]
        
        # Add prediction data (limit text preview for PDF)
        for pred in predictions:
            text_preview = pred['text'][:80] + '...' if len(pred['text']) > 80 else pred['text']
            table_data.append([
                str(pred['id']),
                text_preview,
                pred['prediction'],
                f"{pred['confidence']}%",
                pred['timestamp']
            ])
        
        # Create table
        table = Table(table_data, colWidths=[0.5*inch, 3.5*inch, 1*inch, 1*inch, 1.5*inch])
        
        # Style the table
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Body styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # ID column
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # Prediction column
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # Confidence column
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'predictions_history_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        )
        
    except Exception as e:
        print(f"PDF Export Error: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.errorhandler(404)
def not_found(e):
    """404 error handler"""
    return render_template('404.html'), 404


@app.errorhandler(500)
def internal_error(e):
    """500 error handler"""
    return render_template('500.html'), 500


# Load models when module is imported (works with Flask debug mode reloader)
print("\n" + "="*60)
print("FAKE NEWS DETECTION SYSTEM - WEB APPLICATION")
print("="*60 + "\n")

# Initialize database
database.init_database()

load_models()

print("\n" + "="*60)
print("Starting Flask server...")
print("="*60)
print("\nAccess the application at: http://localhost:5000")
print("Press CTRL+C to stop the server\n")

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000, use_reloader=False)
